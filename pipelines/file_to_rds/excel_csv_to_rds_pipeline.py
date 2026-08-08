# excel_csv_to_rds_pipeline.py
import os
from dotenv import load_dotenv

load_dotenv()
os.environ["DLT_DATA_DIR"] = os.getenv("dlt_pipeline_disk_location")

import dlt
import pandas as pd
import logging
import yaml
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from utils.dot_env_secrets import load_env_as_dict

# Define a specific logger name instead of the root logger
logger = logging.getLogger("data_migration_pipeline")

PIPELINE_NAME = "file_to_rds"

def source_operation_logic(src_operation, merge_key: None) -> dict:
    # Find the operation type
    if src_operation == 'replace':
        # Replace — full reload every run
        operation = {
            "disposition": "replace",
            "strategy": "truncate-and-insert"
        }
    elif src_operation == 'append':
        # Append — add new records each run
        operation = "append"
    elif src_operation == 'merge':
        # Merge — upsert based on primary key
        operation = {
            "disposition": "merge",
            "strategy": "upsert",
            "primary_key": merge_key
        }
    else:
        raise ValueError("The operation type can be replace, append or merge!")

    return operation

def _build_all_text_column_map(columns: list) -> dict:
    """
    Forces EVERY column to dlt's 'text' type — maps to Postgres text/varchar.

    Built from column NAMES (the header row), NOT from data values.
    This guarantees:
      - All-null columns are still created in the target table   (Req 3)
      - No per-column type inference/guessing happens             (Req 2)
      - Schema is 100% deterministic across every file/run
    """
    return {col: {"data_type": "text", "nullable": True} for col in columns}

def _read_csv_as_text(file_path: str, delimiter: str = ",") -> pd.DataFrame:
    """
    CSV has no native types — every value is already just text.
    dtype=str is 100% safe and lossless — satisfies Req 1.
    """
    # Defining delimiter
    if not delimiter:
        delimiter = ","
    
    logger.info(f"Parameters:")
    logger.info(f"   File Path      : {file_path}")
    logger.info(f"   File Format    : CSV")
    logger.info(f"   CSV Delimiter  : {delimiter}")

    df = pd.read_csv(
        file_path,
        dtype=str,
        sep=delimiter,           # configurable delimiter
        keep_default_na=False,   # prevents "NA"/"NULL" strings becoming NaN
        na_values=[""],          # only truly empty cells become NaN/None
    )
    
    return df

def _read_excel_as_text(file_path: str, sheet_name=0) -> pd.DataFrame:
    """
    Reads Excel preserving the value as close to "as displayed" as
    possible — avoids pandas' automatic date/float parsing which can
    alter decimal precision or reformat dates. Satisfies Req 1.
    """
    logger.info(f"Parameters:")
    logger.info(f"   File Path      : {file_path}")
    logger.info(f"   File Format    : EXCEL")

    wb = load_workbook(file_path, data_only=True)
    ws = wb[wb.sheetnames[sheet_name]] if isinstance(sheet_name, int) else wb[sheet_name]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return pd.DataFrame()

    headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]

    data = []
    for row in rows[1:]:
        row_dict = {}
        for col_name, cell_value in zip(headers, row):
            if cell_value is None:
                row_dict[col_name] = None

            elif isinstance(cell_value, bool):
                # MUST come BEFORE int — bool is a subclass of int in Python
                row_dict[col_name] = "TRUE" if cell_value else "FALSE"

            elif isinstance(cell_value, float):
                # Avoid %g — truncates to 6 sig figs and switches to
                # scientific notation for large numbers, corrupting data
                if cell_value == int(cell_value):
                    row_dict[col_name] = str(int(cell_value))
                else:
                    row_dict[col_name] = format(cell_value, ".10f").rstrip("0").rstrip(".")

            elif isinstance(cell_value, int):
                row_dict[col_name] = str(cell_value)

            elif hasattr(cell_value, "isoformat"):
                # datetime/date/time objects — Excel stores these as
                # internal serial numbers, so there is no "original text"
                # to preserve. ISO format is the safest, most deterministic
                # lossless representation (keeps date + time + precision).
                row_dict[col_name] = cell_value.isoformat()

            else:
                row_dict[col_name] = str(cell_value)

        data.append(row_dict)

    df = pd.DataFrame(data, columns=headers)

    return df

def _read_source_file(file_path: str, sheet_name=0, delimiter: str = ",") -> pd.DataFrame:
    """Dispatches to the correct reader based on file extension."""
    file_ext = Path(file_path).suffix.lower()

    if file_ext == ".csv":
        return _read_csv_as_text(file_path, delimiter)
    elif file_ext in (".xlsx", ".xls"):
        return _read_excel_as_text(file_path, sheet_name=sheet_name)
    else:
        raise ValueError(f"Unsupported file type: {file_ext}. Use .csv, .xlsx, or .xls")

def run_excel_csv_to_rds() -> str:
    with open(f"./config/configurations.yaml") as stream:
        # Configurations
        config = yaml.safe_load(stream)
        # Get secret keys & values
        secrets = load_env_as_dict(".env")

        # ── Define your files and configurations from config ──────────────
        files = config.get("file_to_rds")

        # ── success and failures across all files ───────────────────
        successful_tables = []
        failed_tables     = []
        skipped_tables    = []

        # Iterate over each file/table as an individual pipeline
        for table, v in files.items():
            try:
                # Run the pipeline for active status Y
                if v["active_status"].upper() == 'Y':
                    # ── Read config ─────────────────────────────────────────
                    file_path         = v.get("file_path")
                    delimiter         = v.get("delimiter", ",")
                    sheet_name        = v.get("sheet_name", 0)
                    target_rds_schema = v.get("rds_schema_name")
                    batch_size        = v.get("batch_size", 5000)
                    merge_key         = v.get("src_merge_key", None)

                    # ── Destination Connection strings ──────────────────────
                    dest_db_uri = f"{secrets.get(f'pg_connection_string')}/{v.get('rds_db_name')}"
                    # Get source operation
                    operation = source_operation_logic(v.get("src_operation"), merge_key)

                    # ── Validate file_path is defined in YAML ───────────────
                    if not file_path:
                        raise ValueError(f"'file_path' is required in YAML config for '{table}'")

                    logger.info(f"Using FILE extraction for '{table}': {file_path}")

                    # ── Step 1: Read file preserving exact original values ──
                    df = _read_source_file(file_path, sheet_name=sheet_name, delimiter=delimiter)

                    if df.empty:
                        logger.warning(f"File '{file_path}' has no data rows — skipping '{table}'.")
                        skipped_tables.append(table)
                        continue

                    logger.info(f"Read '{file_path}': {len(df):,} rows, {len(df.columns)} columns")
                    logger.info(f"Columns: {list(df.columns)}")

                    # ── Step 2: Build explicit all-text column map from HEADERS
                    # Guarantees Req 2 (all VARCHAR) and Req 3 (null columns kept)
                    text_columns = _build_all_text_column_map(list(df.columns))

                    # ── Step 3: Define dlt resource with forced schema ───────
                    # Factory function captures df/table/columns/batch_size
                    # per loop iteration — avoids late-binding closure bug
                    def _make_file_resource(dataframe, resource_name, columns_map, size):
                        @dlt.resource(name=resource_name, columns=columns_map)
                        def _file_resource():
                            records = dataframe.to_dict(orient="records")
                            for i in range(0, len(records), size):
                                yield records[i: i + size]
                        return _file_resource()

                    source = _make_file_resource(df, table, text_columns, batch_size)

                    # ── Build the pipeline ───────────────────────────────────
                    pipeline = dlt.pipeline(
                        pipeline_name=PIPELINE_NAME,
                        destination=dlt.destinations.postgres(dest_db_uri),
                        dataset_name=target_rds_schema,
                        dev_mode=False
                    )

                    # ── Run pipeline ──────────────────────────────────────
                    logger.info(f"{'='*60}")
                    logger.info(f"Starting Excel/CSV → RDS Pipeline")
                    logger.info(f"Time: {datetime.now()}")
                    logger.info(f"{'='*60}")

                    # ── Run ───────────────────────────────────────────────
                    # write_disposition options:
                    #   "replace" → drops and recreates target table fresh every run
                    #   "append"  → adds rows on top of existing data
                    #   "merge"   → upsert based on primary key (needs primary_key set)
                    load_info = pipeline.run(
                        source,
                        write_disposition=operation,
                    )

                    # ── Check if dlt itself reported any load errors ──────
                    # dlt does not always raise exceptions — sometimes it
                    # captures errors inside load_info. So we check both.
                    if load_info.has_failed_jobs:
                        failed_jobs_detail = [str(job) for job in load_info.failed_jobs]
                        raise RuntimeError(f"dlt reported failed jobs: {failed_jobs_detail}")

                    # ── Print results ─────────────────────────────────────
                    logger.info(f"{'='*60}")
                    logger.info("PIPELINE RESULTS")
                    logger.info(f"{'='*60}")
                    logger.info(load_info)
                    logger.info(
                        f"SOURCE - DESTINATION SUMMARY: {file_path} "
                        f"-> RDS {v.get('rds_db_name')}.{target_rds_schema}.{table}"
                    )
                    logger.info(load_info)
                    successful_tables.append(table)

                # Skip the files having active set as No
                elif v["active_status"].upper() == 'N':
                    msg = "active_status=N — set to Y in configurations to enable"
                    logger.info(f"Skipping '{table}': {msg}")
                    skipped_tables.append(table)
                # Skip the files having incorrect active status
                else:
                    msg = f"Invalid active_status='{v.get('active_status')}' — expected Y or N"
                    logger.warning(f"Skipping '{table}': {msg}")
                    skipped_tables.append(table)

            except ValueError as ve:
                # Specifically catch bad operation type or missing config errors
                logger.error(
                    f"Config error for table '{table}': {ve} — Skipping."
                )
                failed_tables.append({"table": table, "error": str(ve)})
                continue  # move to next file

            except Exception as e:
                # Catch all other errors (file read, connection, schema errors, etc.)
                logger.error(
                    f"Failed to load table '{table}': {type(e).__name__}: {e} — Skipping."
                )
                failed_tables.append({"table": table, "error": str(e)})
                continue  # move to next file

        logging.info('\n')
        # ── Final Summary ─────────────────────────────────────────────────
        logger.info("=" * 60)
        logger.info(f"Pipeline Summary:")
        logger.info(f"   Successful : {len(successful_tables)} table(s) → {successful_tables}")
        logger.info(f"   Skipped    : {len(skipped_tables)} table(s)  → {skipped_tables}")
        logger.info(f"   Failed     : {len(failed_tables)} table(s)")

        for failure in failed_tables:
            logger.info(f"      - {failure['table']}: {failure['error']}")

        logger.info("=" * 60)
        logging.info('\n')

    return "Pipeline Run Completed"
